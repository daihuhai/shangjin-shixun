import os
from datetime import datetime
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook

from .config import EXPORT_DIR, MODEL_DISPLAY_NAME, PLATFORM_NAME
from .db import get_conn, loads, new_id, rows_to_list, utc_now


def _pdf_font_candidates() -> list[Path]:
    windir = Path(os.environ.get("WINDIR", r"C:\Windows"))
    return [
        windir / "Fonts" / "msyh.ttf",
        windir / "Fonts" / "simhei.ttf",
        windir / "Fonts" / "simsun.ttc",
        Path("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/truetype/wqy/wqy-microhei.ttc"),
    ]


def _setup_pdf_font(pdf: FPDF) -> bool:
    for path in _pdf_font_candidates():
        if path.exists():
            pdf.add_font("CJK", "", str(path))
            pdf.set_font("CJK", size=11)
            return True
    pdf.set_font("Helvetica", size=10)
    return False


def _task_filters(task_id: str | None) -> dict:
    return {"task_id": task_id} if task_id else {}


def get_task_detail(task_id: str, user_role: str, user_id: str) -> dict | None:
    with get_conn() as conn:
        task_row = conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if not task_row:
        return None

    filters = {"task_id": task_id}
    if user_role == "student":
        filters["student_id"] = user_id

    rows = build_submission_rows(filters)
    serialized = [serialize_submission(row) for row in rows]
    student_ids = {item["studentId"] for item in serialized}
    pending_checks = [item for item in serialized if item["status"] in {"submitted", "parsed"}]
    pending_scores = [
        item for item in serialized
        if item["status"] in {"checked", "scored"} and not item.get("finalScore")
    ]
    high_risk = [item for item in serialized if item.get("riskLevel") == "高"]
    score_values = [
        item.get("finalScore") or item.get("aiTotalScore")
        for item in serialized
        if item.get("finalScore") or item.get("aiTotalScore")
    ]

    submitters: dict[str, dict] = {}
    for item in serialized:
        sid = item["studentId"]
        if sid not in submitters:
            submitters[sid] = {
                "studentId": sid,
                "studentName": item["studentName"],
                "studentNumber": item["studentNumber"],
                "organization": item["organization"],
                "submissionCount": 0,
                "latestStatus": item["status"],
                "latestScore": item.get("finalScore") or item.get("aiTotalScore"),
                "latestSubmittedAt": item["submittedAt"],
                "latestSubmissionId": item["id"],
                "riskLevel": item.get("riskLevel") or "低",
            }
        submitters[sid]["submissionCount"] += 1

    task = dict(task_row)
    return {
        "task": task,
        "stats": {
            "submissionCount": len(serialized),
            "studentCount": len(student_ids),
            "pendingCheck": len(pending_checks),
            "pendingScore": len(pending_scores),
            "highRisk": len(high_risk),
            "averageScore": round(sum(score_values) / len(score_values), 1) if score_values else None,
        },
        "submitters": sorted(
            submitters.values(),
            key=lambda item: item["latestSubmittedAt"] or "",
            reverse=True,
        ),
        "submissions": serialized,
    }


def build_submission_rows(filters: dict | None = None) -> list[dict]:
    filters = filters or {}
    query = """
        SELECT s.*, t.title AS task_title, t.course, t.class_name,
               u.name AS student_name, u.student_id AS student_number, u.organization,
               sr.ai_total_score, sr.teacher_adjusted_score, sr.final_score,
               pr.status AS parse_status, cr.status AS check_status
        FROM submissions s
        JOIN tasks t ON t.id = s.task_id
        JOIN users u ON u.id = s.student_id
        LEFT JOIN score_records sr ON sr.submission_id = s.id
        LEFT JOIN parse_results pr ON pr.submission_id = s.id
        LEFT JOIN check_reports cr ON cr.submission_id = s.id
        WHERE 1=1
    """
    params = []
    if filters.get("student_id"):
        query += " AND s.student_id = ?"
        params.append(filters["student_id"])
    if filters.get("task_id"):
        query += " AND s.task_id = ?"
        params.append(filters["task_id"])
    if filters.get("organization"):
        # 教师按所属组织过滤：取前2字符作为院系代码匹配下属班级
        # "软件工程学院"(前2字="软件") 匹配 "软件 2301"、"软件2302"
        org = filters["organization"].replace(" ", "")
        dept_code = org[:2] if len(org) >= 2 else org
        query += " AND REPLACE(u.organization, ' ', '') LIKE ?"
        params.append(f"{dept_code}%")
    query += " ORDER BY s.submitted_at DESC"
    with get_conn() as conn:
        rows = conn.execute(query, params).fetchall()
    return rows_to_list(rows)


def dashboard_data(role: str, user_id: str) -> dict:
    submissions = build_submission_rows({"student_id": user_id} if role == "student" else {})
    pending_checks = [item for item in submissions if item["status"] in {"submitted", "parsed"}]
    pending_scores = [item for item in submissions if item["status"] in {"checked", "scored"} and not item.get("final_score")]
    high_risk = [item for item in submissions if item.get("risk_level") == "高"]

    if role == "student":
        return {
            "heroTitle": "学习工作台",
            "metrics": [
                {"label": "进行中任务", "value": str(len({item["task_id"] for item in submissions}) or 1), "state": "进行中"},
                {"label": "我的提交", "value": str(len(submissions)), "state": "稳定"},
                {"label": "待查看反馈", "value": str(len([item for item in submissions if item.get("check_status")])), "state": "待处理"},
                {"label": "平均成绩", "value": _avg_score(submissions), "state": "稳定"}
            ],
            "todos": [
                {
                    "name": item["task_title"],
                    "target": item["course"],
                    "due": item["submitted_at"][:10],
                    "status": item["status"]
                }
                for item in submissions[:5]
            ]
        }

    if role == "teacher":
        return {
            "heroTitle": "教师工作台",
            "cards": [
                {"title": "待核查提交", "value": f"{len(pending_checks)} 份", "note": "需发起智能核查"},
                {"title": "待评分记录", "value": f"{len(pending_scores)} 份", "note": "需教师复核定稿"},
                {"title": "高风险提交", "value": f"{len(high_risk)} 份", "note": "建议优先处理"},
                {"title": "累计提交", "value": f"{len(submissions)} 份", "note": "覆盖全部实训任务"}
            ],
            "warning": {
                "value": f"{len(high_risk)} 份高风险提交待复核",
                "detail": "进入具体任务详情页，在「智能核查」标签中查看证据片段与修复建议。"
            },
            "visits": [
                {"label": "周一", "value": 12 + len(submissions)},
                {"label": "周二", "value": 18 + len(submissions)},
                {"label": "周三", "value": 22 + len(submissions)},
                {"label": "周四", "value": 20 + len(submissions)},
                {"label": "周五", "value": 16 + len(submissions)},
                {"label": "周六", "value": 8},
                {"label": "周日", "value": 6}
            ],
            "todos": [
                {
                    "name": item["task_title"],
                    "target": item["student_name"],
                    "due": item["submitted_at"][:10],
                    "status": item["status"]
                }
                for item in (pending_checks + pending_scores)[:5]
            ]
        }

    with get_conn() as conn:
        user_count = conn.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
        task_count = conn.execute("SELECT COUNT(*) AS c FROM tasks").fetchone()["c"]
        log_count = conn.execute("SELECT COUNT(*) AS c FROM model_call_logs WHERE success = 1").fetchone()["c"]
        export_count = conn.execute("SELECT COUNT(*) AS c FROM report_snapshots").fetchone()["c"]

    return {
        "heroTitle": "管理后台",
        "metrics": [
            {"label": "平台用户", "value": str(user_count), "state": "稳定"},
            {"label": "实训任务", "value": str(task_count), "state": "稳定"},
            {"label": "模型成功调用", "value": str(log_count), "state": "稳定"},
            {"label": "报表导出次数", "value": str(export_count), "state": "稳定"}
        ],
        "charts": {
            "course": [
                {"label": "待核查", "value": len(pending_checks)},
                {"label": "待评分", "value": len(pending_scores)},
                {"label": "高风险", "value": len(high_risk)}
            ]
        },
        "rankings": [
            {"name": "尚进大模型", "indicator": MODEL_DISPLAY_NAME, "status": "在线"},
            {"name": "Java Web 实训", "indicator": f"{len(submissions)} 份提交", "status": "活跃"}
        ]
    }


def _avg_score(submissions: list[dict]) -> str:
    scores = [item["final_score"] or item["teacher_adjusted_score"] or item["ai_total_score"] for item in submissions]
    scores = [score for score in scores if score is not None]
    if not scores:
        return "--"
    return f"{sum(scores) / len(scores):.1f}"


def export_excel(report_type: str, created_by: str, task_id: str | None = None, student_id: str | None = None) -> dict:
    filters = _task_filters(task_id)
    if student_id:
        filters["student_id"] = student_id
    rows = build_submission_rows(filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "实训评价"
    ws.append(["学生", "学号", "班级", "任务", "提交时间", "状态", "风险", "尚进大模型分", "教师分", "最终分"])
    for row in rows:
        ws.append([
            row["student_name"],
            row["student_number"],
            row["organization"],
            row["task_title"],
            row["submitted_at"][:19].replace("T", " "),
            row["status"],
            row["risk_level"],
            _row_get(row, "ai_total_score"),
            _row_get(row, "teacher_adjusted_score"),
            _row_get(row, "final_score")
        ])

    snapshot_id = new_id()[:8]
    filename = f"{report_type}-{snapshot_id}.xlsx"
    file_path = EXPORT_DIR / filename
    wb.save(file_path)
    _save_snapshot(report_type, "excel", filename, str(file_path), created_by)
    return {"filename": filename, "snapshotId": snapshot_id, "downloadUrl": f"/api/reports/download/{filename}"}


def export_pdf(report_type: str, created_by: str, task_id: str | None = None, student_id: str | None = None) -> dict:
    filters = _task_filters(task_id)
    if student_id:
        filters["student_id"] = student_id
    rows = build_submission_rows(filters)
    pdf = FPDF()
    pdf.add_page()
    has_cjk = _setup_pdf_font(pdf)
    title = f"{PLATFORM_NAME} 实训评价报表" if has_cjk else "Training Evaluation Report"
    pdf.set_font("CJK" if has_cjk else "Helvetica", size=14)
    pdf.cell(0, 10, title, ln=1)
    pdf.set_font("CJK" if has_cjk else "Helvetica", size=10)
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pdf.cell(0, 8, f"生成时间：{generated}" if has_cjk else f"Generated: {generated}", ln=1)
    if task_id and rows:
        task_title = rows[0]["task_title"]
        pdf.cell(0, 8, f"任务：{task_title}" if has_cjk else f"Task: {task_title}", ln=1)
    pdf.ln(4)

    for row in rows[:50]:
        if has_cjk:
            line = (
                f"{row['student_name']}（{row['student_number']}）| "
                f"状态={row['status']} | 风险={row['risk_level']} | "
                f"尚进大模型分={row.get('ai_total_score') or '--'} | 最终分={row.get('final_score') or '--'}"
            )
        else:
            line = (
                f"{row['student_number']} | status={row['status']} | "
                f"risk={row['risk_level']} | final={row.get('final_score') or '--'}"
            )
        pdf.multi_cell(0, 6, line)
        pdf.ln(1)

    snapshot_id = new_id()[:8]
    filename = f"{report_type}-{snapshot_id}.pdf"
    file_path = EXPORT_DIR / filename
    pdf.output(str(file_path))
    _save_snapshot(report_type, "pdf", filename, str(file_path), created_by)
    return {"filename": filename, "snapshotId": snapshot_id, "downloadUrl": f"/api/reports/download/{filename}"}


def report_summary(task_id: str | None = None, user_role: str = "", user_id: str = "", organization: str = "") -> dict:
    filters = _task_filters(task_id)
    if user_role == "student":
        filters["student_id"] = user_id
    if organization and user_role != "student":
        filters["organization"] = organization
    rows = build_submission_rows(filters)
    score_values = [row.get("final_score") or row.get("ai_total_score") for row in rows if row.get("final_score") or row.get("ai_total_score")]
    distribution = {"90-100": 0, "80-89": 0, "70-79": 0, "60-69": 0, "<60": 0}
    for score in score_values:
        if score >= 90:
            distribution["90-100"] += 1
        elif score >= 80:
            distribution["80-89"] += 1
        elif score >= 70:
            distribution["70-79"] += 1
        elif score >= 60:
            distribution["60-69"] += 1
        else:
            distribution["<60"] += 1

    risk_stats = {"低": 0, "中": 0, "高": 0}
    for row in rows:
        risk_stats[row.get("risk_level") or "低"] = risk_stats.get(row.get("risk_level") or "低", 0) + 1

    return {
        "totalSubmissions": len(rows),
        "averageScore": round(sum(score_values) / len(score_values), 1) if score_values else 0,
        "distribution": [{"label": key, "value": value} for key, value in distribution.items()],
        "riskStats": [{"label": key, "value": value} for key, value in risk_stats.items()],
        "recentExports": _recent_exports(user_id)
    }


def _recent_exports(user_id: str = "") -> list[dict]:
    with get_conn() as conn:
        if user_id:
            rows = conn.execute(
                "SELECT id, report_type, format, filename, created_at FROM report_snapshots WHERE created_by = ? ORDER BY created_at DESC LIMIT 8",
                (user_id,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, report_type, format, filename, created_at FROM report_snapshots ORDER BY created_at DESC LIMIT 8"
            ).fetchall()
    return [
        {
            "id": row["id"],
            "type": row["report_type"],
            "format": row["format"],
            "filename": row["filename"],
            "createdAt": row["created_at"],
            "downloadUrl": f"/api/reports/download/{row['filename']}"
        }
        for row in rows
    ]


def _save_snapshot(report_type: str, fmt: str, filename: str, file_path: str, created_by: str):
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO report_snapshots (id, report_type, format, filename, file_path, created_by, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (new_id(), report_type, fmt, filename, file_path, created_by, utc_now())
        )


def _row_get(row, key, default=None):
    """安全获取 sqlite3.Row 的值，兼容 .get() 语义"""
    try:
        return row[key]
    except (IndexError, KeyError):
        return default


def _parse_score_dimensions(dimensions_json: str | None) -> dict:
    """
    解析评分维度数据，兼容新旧两种格式
    - 旧格式：直接是 dimensions 数组
    - 新格式：包含 dimensions + feedback 信息的对象
    """
    if not dimensions_json:
        return {"dimensions": []}

    data = loads(dimensions_json, {})

    # 新版格式（包含 feedback, highlights, errors 等）
    if isinstance(data, dict) and "dimensions" in data:
        return {
            "dimensions": data.get("dimensions", []),
            "feedback": data.get("feedback", ""),
            "summary": data.get("summary", ""),
            "highlights": data.get("highlights", []),
            "errors": data.get("errors", []),
            "similarityScore": data.get("similarityScore"),
            "correctnessScore": data.get("correctnessScore"),
            "completenessScore": data.get("completenessScore"),
        }

    # 旧版格式（直接是数组）
    if isinstance(data, list):
        return {"dimensions": data}

    # 兜底
    return {"dimensions": []}


def serialize_submission(row: dict) -> dict:
    with get_conn() as conn:
        files = rows_to_list(
            conn.execute(
                "SELECT id, filename, file_type, file_size, created_at FROM submission_files WHERE submission_id = ?",
                (row["id"],)
            ).fetchall()
        )
        parse_row = conn.execute("SELECT * FROM parse_results WHERE submission_id = ?", (row["id"],)).fetchone()
        check_row = conn.execute("SELECT * FROM check_reports WHERE submission_id = ?", (row["id"],)).fetchone()
        score_row = conn.execute("SELECT * FROM score_records WHERE submission_id = ?", (row["id"],)).fetchone()
        items = []
        if check_row:
            items = rows_to_list(
                conn.execute(
                    "SELECT * FROM check_items WHERE report_id = ? ORDER BY sort_order ASC",
                    (check_row["id"],)
                ).fetchall()
            )

    return {
        "id": row["id"],
        "taskId": row["task_id"],
        "taskTitle": _row_get(row, "task_title") or _row_get(row, "title"),
        "course": _row_get(row, "course"),
        "className": _row_get(row, "class_name"),
        "studentId": row["student_id"],
        "studentName": _row_get(row, "student_name"),
        "studentNumber": _row_get(row, "student_number"),
        "organization": _row_get(row, "organization"),
        "version": row["version"],
        "status": row["status"],
        "riskLevel": _row_get(row, "risk_level") or "低",
        "remark": _row_get(row, "remark"),
        "evaluationError": _row_get(row, "evaluation_error"),
        "submittedAt": row["submitted_at"],
        "files": files,
        "parseResult": dict(parse_row) if parse_row else None,
        "checkReport": {
            **dict(check_row),
            "items": [
                {
                    "id": item["id"],
                    "name": item["name"],
                    "category": item["category"],
                    "conclusion": item["conclusion"],
                    "riskLevel": item["risk_level"],
                    "evidence": item["evidence"],
                    "suggestion": item["suggestion"],
                    "needsReview": bool(item["needs_review"]),
                    "teacherMark": item["teacher_mark"]
                }
                for item in items
            ],
            "effectiveHighRiskCount": sum(1 for item in items if item["risk_level"] == "高" and item.get("teacher_mark") != "误判")
        } if check_row else None,
        "scoreRecord": {
            "ai_total_score": score_row["ai_total_score"],
            "teacher_adjusted_score": score_row["teacher_adjusted_score"],
            "final_score": score_row["final_score"],
            "adjustment_reason": score_row["adjustment_reason"],
            "teacher_comment": score_row["teacher_comment"],
            **_parse_score_dimensions(score_row["dimensions_json"])
        } if score_row else None,
        "aiTotalScore": _row_get(row, "ai_total_score"),
        "teacherAdjustedScore": _row_get(row, "teacher_adjusted_score"),
        "finalScore": _row_get(row, "final_score")
    }


def teacher_profile(user_id: str) -> dict:
    """生成教师教学画像：教学能力维度 + 标签 + 基础数据（不含AI建议）"""
    with get_conn() as conn:
        # 该教师创建的课程
        courses = rows_to_list(conn.execute(
            "SELECT id, name FROM courses WHERE created_by = ?", (user_id,)
        ).fetchall())
        course_ids = [c["id"] for c in courses]

        # 该教师创建的任务（含 course_id 匹配 + course 名称匹配）
        if course_ids:
            placeholders = ",".join("?" for _ in course_ids)
            tasks_by_id = rows_to_list(conn.execute(
                f"SELECT id, title, course_id, course FROM tasks WHERE course_id IN ({placeholders})", course_ids
            ).fetchall())
        else:
            tasks_by_id = []
        # 也查一下没有 course_id 但 course 名称匹配的任务
        task_titles_set = set()
        for c in courses:
            cname = c["name"].replace(" ", "")
            extra = rows_to_list(conn.execute(
                "SELECT id, title, course_id, course FROM tasks WHERE course_id IS NULL AND REPLACE(course, ' ', '') = ?",
                (cname,)
            ).fetchall())
            for t in extra:
                if t["id"] not in [x["id"] for x in tasks_by_id]:
                    tasks_by_id.append(t)

        task_ids = [t["id"] for t in tasks_by_id]
        total_tasks = len(tasks_by_id)

        # 收集这些任务的所有学生提交
        submissions = []
        if task_ids:
            placeholders = ",".join("?" for _ in task_ids)
            submissions = rows_to_list(conn.execute(
                f"""SELECT s.*, t.title AS task_title, t.course,
                       u.name AS student_name, u.organization AS student_org,
                       sr.ai_total_score, sr.teacher_adjusted_score, sr.final_score,
                       sr.dimensions_json, sr.teacher_comment,
                       pr.status AS parse_status, cr.status AS check_status,
                       s.risk_level
                FROM submissions s
                JOIN tasks t ON t.id = s.task_id
                JOIN users u ON u.id = s.student_id
                LEFT JOIN score_records sr ON sr.submission_id = s.id
                LEFT JOIN parse_results pr ON pr.submission_id = s.id
                LEFT JOIN check_reports cr ON cr.submission_id = s.id
                WHERE s.task_id IN ({placeholders})
                ORDER BY s.submitted_at DESC""",
                task_ids
            ).fetchall())

        # 统计指标
        total_submissions = len(submissions)
        unique_students = set(s["student_id"] for s in submissions)
        unique_student_count = len(unique_students)

        # 评分统计
        all_scores = []
        scored_count = 0
        pending_review = 0
        high_risk_count = 0
        similarity_scores = []
        correctness_scores = []
        completeness_scores = []
        feedback_count = 0
        teacher_adjusted_count = 0
        task_sub_count = {}  # task_id -> submission count

        for row in submissions:
            tid = row.get("task_id")
            task_sub_count[tid] = task_sub_count.get(tid, 0) + 1

            score_row = None
            with get_conn() as conn2:
                score_row = conn2.execute(
                    "SELECT * FROM score_records WHERE submission_id = ?", (row["id"],)
                ).fetchone()

            if score_row:
                final = row.get("final_score") or _row_get(score_row, "ai_total_score")
                if final is not None:
                    all_scores.append(final)
                    scored_count += 1
                if _row_get(score_row, "teacher_adjusted_score") is not None:
                    teacher_adjusted_count += 1
                if _row_get(score_row, "teacher_comment"):
                    feedback_count += 1

                dims = _parse_score_dimensions(_row_get(score_row, "dimensions_json"))
                sim = dims.get("similarityScore")
                cor = dims.get("correctnessScore")
                com = dims.get("completenessScore")
                if sim is None or cor is None or com is None:
                    for dim in dims.get("dimensions", []):
                        dim_name = dim.get("name", "")
                        dim_score = dim.get("aiScore") or dim.get("score")
                        if dim_score is not None:
                            if "代码质量" in dim_name or "相似度" in dim_name:
                                if sim is None: sim = dim_score
                            elif "文档" in dim_name or "正确性" in dim_name:
                                if cor is None: cor = dim_score
                            elif "功能" in dim_name or "完整" in dim_name:
                                if com is None: com = dim_score
                if sim is not None: similarity_scores.append(sim)
                if cor is not None: correctness_scores.append(cor)
                if com is not None: completeness_scores.append(com)

            if row.get("risk_level") == "高":
                high_risk_count += 1

            status = row.get("status") or ""
            if status in ("submitted", "parsed"):
                pending_review += 1

        avg_score = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0
        score_rate = round(scored_count / max(total_submissions, 1) * 100, 1)
        feedback_rate = round(feedback_count / max(scored_count, 1) * 100, 1)

        # 教学能力维度（基于教师教学数据计算）
        # 任务设计力：该教师创建的任务数 / 平台平均任务数（归一化）
        with get_conn() as conn3:
            platform_task_count = conn3.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]
            platform_teacher_count = conn3.execute(
                "SELECT COUNT(DISTINCT created_by) FROM courses WHERE created_by IN (SELECT id FROM users WHERE role='teacher')"
            ).fetchone()[0]
        avg_tasks_per_teacher = platform_task_count / max(platform_teacher_count, 1)
        task_design_rate = min(100, round(total_tasks / max(avg_tasks_per_teacher, 1) * 50, 1))

        # 评分效率：已评分数 / 总提交数
        grading_rate = score_rate

        # 学生覆盖度：参与学生数 / 平台总学生数
        with get_conn() as conn4:
            platform_student_count = conn4.execute(
                "SELECT COUNT(*) FROM users WHERE role='student'"
            ).fetchone()[0]
        coverage_rate = min(100, round(unique_student_count / max(platform_student_count, 1) * 100, 1)) if platform_student_count > 0 else 0

        # 反馈深度：有教师评语的比例
        feedback_depth_rate = feedback_rate

        # 综合评分 = 四个维度的加权平均
        overall_score = round(
            task_design_rate * 0.25 + grading_rate * 0.30 + coverage_rate * 0.20 + feedback_depth_rate * 0.25,
            1
        )

        # 能力维度（归一化到百分比用于饼图）
        total_rate = task_design_rate + grading_rate + coverage_rate + feedback_depth_rate
        if total_rate > 0:
            abilities = [
                {"name": "任务设计力", "value": round(task_design_rate / total_rate * 100), "rate": task_design_rate},
                {"name": "评分效率", "value": round(grading_rate / total_rate * 100), "rate": grading_rate},
                {"name": "学生覆盖度", "value": round(coverage_rate / total_rate * 100), "rate": coverage_rate},
                {"name": "反馈深度", "value": round(feedback_depth_rate / total_rate * 100), "rate": feedback_depth_rate},
            ]
        else:
            abilities = [
                {"name": "任务设计力", "value": 25, "rate": 0},
                {"name": "评分效率", "value": 25, "rate": 0},
                {"name": "学生覆盖度", "value": 25, "rate": 0},
                {"name": "反馈深度", "value": 25, "rate": 0},
            ]

        # 综合评级
        if overall_score >= 85:
            grade_label, grade_desc = "A", "优秀"
        elif overall_score >= 70:
            grade_label, grade_desc = "B+", "良好"
        elif overall_score >= 60:
            grade_label, grade_desc = "B", "及格"
        elif overall_score >= 40:
            grade_label, grade_desc = "C+", "需努力"
        else:
            grade_label, grade_desc = "C", "待提升"

        # 标签推断
        dim_sorted = sorted(abilities, key=lambda x: x["rate"], reverse=True)
        strength = dim_sorted[0]["name"] if abilities else "数据不足"
        weak = dim_sorted[-1]["name"] if abilities else "数据不足"

        if grading_rate >= 80 and feedback_depth_rate >= 60:
            learn_type = "严谨细致型"
        elif task_design_rate >= 70:
            learn_type = "创新设计型"
        elif coverage_rate >= 60:
            learn_type = "广泛覆盖型"
        else:
            learn_type = "稳步发展型"

        if grading_rate > feedback_depth_rate + 20:
            trend = "评分积极"
        elif feedback_depth_rate > grading_rate + 20:
            trend = "反馈深入"
        else:
            trend = "均衡发展"

        recommend = f"加强{weak}方面的教学投入" if weak != "数据不足" else "保持当前教学节奏"

        return {
            "gradeLabel": grade_label,
            "gradeDesc": grade_desc,
            "avgScore": avg_score,
            "totalSubmissions": total_submissions,
            "abilities": abilities,
            "tags": {
                "learnType": learn_type,
                "strength": strength,
                "weakness": weak,
                "trend": trend,
                "recommend": recommend,
            },
            "llmContext": {
                "courseCount": len(courses),
                "taskCount": total_tasks,
                "submissionCount": total_submissions,
                "studentCount": unique_student_count,
                "scoredCount": scored_count,
                "avgScore": avg_score,
                "scoreRate": score_rate,
                "feedbackRate": feedback_rate,
                "highRiskCount": high_risk_count,
                "pendingReview": pending_review,
                "teacherAdjustedCount": teacher_adjusted_count,
                "taskTitles": list(set(t.get("title", "") for t in tasks_by_id))[:10],
                "courseNames": [c["name"] for c in courses],
            },
        }


def student_profile(user_id: str) -> dict:
    """生成学生画像：能力分布饼图 + 标签 + 基础数据（不含AI建议，建议由路由层异步调用LLM）"""
    rows = build_submission_rows({"student_id": user_id})

    # 收集所有评分记录的维度分数
    similarity_scores = []
    correctness_scores = []
    completeness_scores = []
    all_scores = []
    task_titles = []
    feedbacks = []
    highlights_set = set()
    errors_list = []
    scored_task_ids = set()  # 记录有评分的任务ID（去重）

    for row in rows:
        score_row = None
        with get_conn() as conn:
            score_row = conn.execute(
                "SELECT * FROM score_records WHERE submission_id = ?", (row["id"],)
            ).fetchone()
        if not score_row:
            continue

        dims = _parse_score_dimensions(_row_get(score_row, "dimensions_json"))
        
        # 尝试从新格式获取分数
        sim = dims.get("similarityScore")
        cor = dims.get("correctnessScore")
        com = dims.get("completenessScore")
        
        # 如果新格式没有数据，尝试从旧格式维度数组中提取
        if sim is None or cor is None or com is None:
            for dim in dims.get("dimensions", []):
                dim_name = dim.get("name", "")
                dim_score = dim.get("aiScore") or dim.get("score")
                if dim_score is not None:
                    if "代码质量" in dim_name or "相似度" in dim_name:
                        if sim is None:
                            sim = dim_score
                    elif "文档" in dim_name or "正确性" in dim_name:
                        if cor is None:
                            cor = dim_score
                    elif "功能" in dim_name or "完整" in dim_name:
                        if com is None:
                            com = dim_score

        if sim is not None:
            similarity_scores.append(sim)
        if cor is not None:
            correctness_scores.append(cor)
        if com is not None:
            completeness_scores.append(com)

        final = row.get("final_score") or _row_get(score_row, "ai_total_score")
        if final is not None:
            all_scores.append(final)
            scored_task_ids.add(_row_get(row, "task_id"))  # 只记录有评分的任务

        task_titles.append(_row_get(row, "task_title") or _row_get(row, "title") or "未知任务")
        if dims.get("feedback"):
            feedbacks.append(dims["feedback"])
        for h in dims.get("highlights", []):
            highlights_set.add(h)
        for e in dims.get("errors", []):
            errors_list.append(e)

    n = len(all_scores)
    avg_score = round(sum(all_scores) / n, 1) if all_scores else 0

    # 获取学生的课程总任务数（通过学生的提交记录关联到课程）
    total_tasks_in_course = 0
    if rows:
        first_row = rows[0]
        task_id = _row_get(first_row, "task_id")
        with get_conn() as conn:
            course_info = conn.execute(
                "SELECT course_id, course FROM tasks WHERE id = ?", (task_id,)
            ).fetchone()
            if course_info:
                course_id = _row_get(course_info, "course_id")
                course_name = _row_get(course_info, "course")
                # 获取该课程的总任务数
                if course_id:
                    total_tasks_in_course = conn.execute(
                        "SELECT COUNT(*) FROM tasks WHERE course_id = ?", (course_id,)
                    ).fetchone()[0]
                elif course_name:
                    # 旧版任务可能没有course_id，用course名称匹配（去除空格）
                    course_name_clean = course_name.replace(" ", "")
                    total_tasks_in_course = conn.execute(
                        "SELECT COUNT(*) FROM tasks WHERE REPLACE(course, ' ', '') = ?", (course_name_clean,)
                    ).fetchone()[0]

    # 能力维度计算（基于评分维度的均值，转换为百分制得分率）
    # similarityScore: 0-50分 -> 0-100%（实践能力）
    sim_rate = round(sum(similarity_scores) / len(similarity_scores) / 50 * 100, 1) if similarity_scores else 0
    # correctnessScore: 0-30分 -> 0-100%（理论掌握）
    cor_rate = round(sum(correctness_scores) / len(correctness_scores) / 30 * 100, 1) if correctness_scores else 0
    # completenessScore: 0-20分 -> 0-100%（项目完成度）
    com_rate = round(sum(completeness_scores) / len(completeness_scores) / 20 * 100, 1) if completeness_scores else 0
    # 出勤与参与度 = 完成的不同任务数 / 课程总任务数 × 100%（上限100%）
    completed_unique_tasks = len(scored_task_ids)
    attendance_rate = min(100, round(completed_unique_tasks / max(total_tasks_in_course, 1) * 100, 1)) if total_tasks_in_course > 0 else 0

    # 归一化到百分比用于饼图显示（总和=100%）
    total_rate = sim_rate + cor_rate + com_rate + attendance_rate
    if total_rate > 0:
        abilities = [
            {"name": "实践能力", "value": round(sim_rate / total_rate * 100), "rate": sim_rate, "maxScore": 50},
            {"name": "理论掌握", "value": round(cor_rate / total_rate * 100), "rate": cor_rate, "maxScore": 30},
            {"name": "项目完成度", "value": round(com_rate / total_rate * 100), "rate": com_rate, "maxScore": 20},
            {"name": "出勤与参与", "value": round(attendance_rate / total_rate * 100), "rate": attendance_rate, "maxScore": total_tasks_in_course},
        ]
    else:
        # 无数据时显示0%
        abilities = [
            {"name": "实践能力", "value": 0, "rate": 0, "maxScore": 50},
            {"name": "理论掌握", "value": 0, "rate": 0, "maxScore": 30},
            {"name": "项目完成度", "value": 0, "rate": 0, "maxScore": 20},
            {"name": "出勤与参与", "value": 0, "rate": 0, "maxScore": total_tasks_in_course or 1},
        ]

    # 综合评级
    pct = avg_score if avg_score > 0 else 0
    if pct >= 85:
        grade_label, grade_desc = "A", "优秀"
    elif pct >= 70:
        grade_label, grade_desc = "B+", "良好"
    elif pct >= 60:
        grade_label, grade_desc = "B", "及格"
    elif pct >= 40:
        grade_label, grade_desc = "C+", "需努力"
    else:
        grade_label, grade_desc = "C", "待提升"

    # 标签推断（基于得分率）
    if sim_rate >= 70 and cor_rate < 60:
        learn_type = "实践驱动型"
    elif cor_rate >= 75:
        learn_type = "理论扎实型"
    elif com_rate >= 70:
        learn_type = "完整规范型"
    else:
        learn_type = "均衡发展型"

    # 优势能力（基于得分率排序）
    dim_sorted = sorted(abilities, key=lambda x: x["rate"], reverse=True)
    strength = dim_sorted[0]["name"] if abilities else "数据不足"

    # 薄弱环节
    weak = dim_sorted[-1]["name"] if abilities else "数据不足"

    # 成长趋势（基于最近提交 vs 平均）
    if len(all_scores) >= 2:
        recent_avg = sum(all_scores[:min(3, len(all_scores))]) / min(3, len(all_scores))
        overall_avg = sum(all_scores) / len(all_scores)
        if recent_avg > overall_avg + 5:
            trend = "稳步上升"
        elif recent_avg < overall_avg - 5:
            trend = "需要加油"
        else:
            trend = "保持稳定"
    else:
        trend = "数据不足" if n == 0 else "起步阶段"

    # 推荐方向
    if weak == "理论掌握":
        recommend = "加强理论基础学习"
    elif weak == "实践能力":
        recommend = "多动手做项目实战"
    elif weak == "项目完成度":
        recommend = "注重代码完整性"
    else:
        recommend = "保持积极参与"

    return {
        "gradeLabel": grade_label,
        "gradeDesc": grade_desc,
        "avgScore": avg_score,
        "totalSubmissions": n,
        "abilities": abilities,
        "tags": {
            "learnType": learn_type,
            "strength": strength,
            "weakness": weak,
            "trend": trend,
            "recommend": recommend,
        },
        # 用于AI建议的上下文数据
        "llmContext": {
            "taskCount": total_tasks_in_course,
            "completedTasks": completed_unique_tasks,
            "submissionCount": n,
            "avgScore": avg_score,
            "taskTitles": list(set(task_titles)),
            "highlights": list(highlights_set)[:10],
            "errorCount": len(errors_list),
            "recentScores": all_scores[:5],
        },
    }
